from DataBase import Database_test
from openpyxl  import load_workbook
import xlrd
import shutil


with Database_test() as data_t:
    count = data_t.nic_count
    hosts = data_t.gen_install_success()

ipmi_ips = [i[0] for i in hosts]
print(ipmi_ips)
excel_path = 'base.xlsx'
shutil.copyfile(excel_path, 'base_new.xlsx')
excel = xlrd.open_workbook(excel_path, encoding_override="utf-8")
workbooknew = load_workbook('base_new.xlsx')
sheets = excel.sheet_names()
table = excel.sheet_by_name(sheets[0])
sheets_copy = workbooknew['Sheet1']
rows_num = table.nrows

for i in range(1, rows_num):
    switch_index = 11 + count
    init_info = table.row_values(i)
    ipmi_ip = init_info[1]
    init_mac = init_info[10:10 + count]
    init_macs = [m.lower() for m in init_mac]
    init_switchs = init_info[switch_index:switch_index + count]
    index_new = ipmi_ips.index(ipmi_ip)
    new_macs = hosts[index_new][1:1 + count]
    if init_macs != new_macs:
        print('%s Port order error' % ipmi_ip)
        new_switchs = []
        for id in range(len(new_macs)):
            ind = init_macs.index(new_macs[id])
            new_switchs.append(init_switchs[id])
            sheets_copy.cell(i+1, switch_index + ind + 1).value = init_switchs[id]
            print('%s %s %s' % (i+1, switch_index + ind + 1, init_switchs[id]))
    else:
        continue

workbooknew.save('base_new.xlsx')
