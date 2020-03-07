# -*- coding:utf-8 -*-
import xlrd
import copy
import shutil
from configparser import ConfigParser
from openpyxl import load_workbook
from database import Database_test

cp = ConfigParser()
cp.read("bms.ini")

EXCEL_PATH = cp.get("excel", "path")

with Database_test() as data_t:
    count = data_t.nic_count
    hosts = data_t.select_info('host', '*')

ipmi_ips = [i[0] for i in hosts]

excel_path = EXCEL_PATH
shutil.copyfile(excel_path, 'base_new.xlsx')
excel = xlrd.open_workbook(excel_path, encoding_override="utf-8")
workbooknew = load_workbook('base_new.xlsx')
sheets = excel.sheet_names()
table = excel.sheet_by_name(sheets[0])
sheets_copy = workbooknew['Sheet1']
rows_num = table.nrows

for i in range(1, rows_num):
    switch_index = 11 + count
    init_info = table.row_values(i)
    ipmi_ip = init_info[1]
    init_macs = init_info[10:10 + count]
    init_switchs = init_info[switch_index:switch_index + count]
    index_new = ipmi_ips.index(ipmi_ip)
    new_mac = hosts[index_new][1:1 + count]
    new_macs = [m.upper() for m in new_mac]

    new_switchs = []
    mac_res = [m1 for m1 in init_macs if m1 in new_macs]
    if len(mac_res) != len(new_macs):
        for mac_id in range(len(new_macs)):
            sheets_copy.cell(i + 1, 10 + mac_id + 1).value = new_macs[mac_id]
        with open('excel_log', 'a') as fp:
            fp.write('%s mac address error' % ipmi_ip)
        init_macs = new_macs

    with Database_test() as data_t:
        dhcp_mac = data_t.select_info('create_bms', 'dhcp_mac', ipmi_ip=ipmi_ip)
        dhcp_mac = dhcp_mac[0]
        print(dhcp_mac)
    new_macs1 = copy.deepcopy(new_macs)
    if new_macs1[0] != dhcp_mac:
        print('%s Port order error' % ipmi_ip)
        ind = new_macs1.index(dhcp_mac)
        new_macs1[0], new_macs1[ind] = new_macs1[ind], new_macs1[0]
        for id in range(len(new_macs1)):
            ind = init_macs.index(new_macs1[id])
            new_switchs.append(init_switchs[id])
            print('%s %s %s' % (i + 1, switch_index + ind + 1, init_switchs[id]))
            sheets_copy.cell(i+1, switch_index + ind + 1).value = init_switchs[id]
    else:
        continue

workbooknew.save('base_new.xlsx')
